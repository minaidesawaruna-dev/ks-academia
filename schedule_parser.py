"""Parser for the KS Academia teacher schedule workbooks.

The workbooks are calendar grids rather than tables:

* Each worksheet is one month (``Aug``, ``Sep``, ``July`` ...).
* One header row holds day labels such as ``15(SAT)``.  A day label is
  usually merged over two columns: the first column holds the class cell,
  the second holds status cells (Online / Recording / Cancelled).
* One or more columns hold the 30 minute time axis and are repeated across
  the sheet purely for readability.
* A class cell looks like::

      G11 Chem HL B
      9am-11am
      Nam Jihoon
      Oh Minseok
      ...

* A status cell looks like::

      Recording
      Yoon Chaewon
      Nam Jihoon

  and may contain several labelled sections in one cell.

The public API is intentionally small:

``get_sheet_names(source)``
``infer_month(sheet_name)``
``parse_schedule(source, sheet_name, year)``
``parse_workbook(source, sheet_names, year)``
"""

from __future__ import annotations

import calendar
import datetime as dt
import difflib
import io
import re
from collections import Counter, defaultdict
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

__all__ = [
    "get_sheet_names",
    "canonicalise_names",
    "infer_month",
    "parse_schedule",
    "parse_workbook",
    "ATTENDING",
    "ONLINE",
    "RECORDING",
    "CANCELLED",
]

# --------------------------------------------------------------------------
# Status vocabulary
# --------------------------------------------------------------------------

ATTENDING = "Attending"
ONLINE = "Online"
RECORDING = "Recording"
CANCELLED = "Cancelled"

# Order matters: "cancel" is checked first so that a line such as
# "Cancelled online class" is treated as a cancellation.
STATUS_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    (CANCELLED, re.compile(r"cancel+ed|cancell?ed|cancel|취소|결석", re.I)),
    (ONLINE, re.compile(r"on\s*-?\s*line|zoom|온라인", re.I)),
    (RECORDING, re.compile(r"record(?:ing|ed)?|녹화|영상", re.I)),
)

# Words that may sit beside a status keyword without being a student name.
_LABEL_FILLER = {
    "review",
    "only",
    "lesson",
    "class",
    "student",
    "students",
    "all",
    "수업",
    "학생",
}

# --------------------------------------------------------------------------
# Regular expressions
# --------------------------------------------------------------------------

DAY_HEADER_RE = re.compile(
    r"^\s*(?P<day>\d{1,2})\s*[\(\[]?\s*(?P<dow>[A-Za-z]{2,9})?\.?\s*[\)\]]?\s*$"
)

TIME_RANGE_RE = re.compile(
    r"(?P<sh>\d{1,2})\s*(?:[.:]\s*(?P<sm>\d{2}))?\s*(?P<sap>[ap]\.?\s?m\.?)?"
    r"\s*[-~\u2010-\u2015]\s*"
    r"(?P<eh>\d{1,2})\s*(?:[.:]\s*(?P<em>\d{2}))?\s*(?P<eap>[ap]\.?\s?m\.?)?",
    re.I,
)

WEEKDAYS = {
    "mon": 0, "monday": 0,
    "tue": 1, "tues": 1, "tuesday": 1,
    "wed": 2, "weds": 2, "wednesday": 2,
    "thu": 3, "thur": 3, "thurs": 3, "thursday": 3,
    "fri": 4, "friday": 4,
    "sat": 5, "saturday": 5,
    "sun": 6, "sunday": 6,
}

# Zero-width and non-breaking characters that survive copy-paste into Excel and
# make two visually identical names compare as different strings.
INVISIBLE = dict.fromkeys(
    map(ord, "\u2060\u200b\u200c\u200d\ufeff\u00a0"), " "
)

MERGE_THRESHOLD = 0.88


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------


def _as_stream(source: Any) -> io.BytesIO:
    """Accept raw bytes, a path, or a file-like object."""
    if isinstance(source, (bytes, bytearray)):
        return io.BytesIO(bytes(source))
    if hasattr(source, "read"):
        data = source.read()
        if hasattr(source, "seek"):
            try:
                source.seek(0)
            except Exception:  # pragma: no cover - non seekable stream
                pass
        return io.BytesIO(data)
    with open(source, "rb") as handle:
        return io.BytesIO(handle.read())


def _cell_lines(value: Any) -> list[str]:
    if not isinstance(value, str):
        return []
    return [line.strip() for line in value.splitlines() if line.strip()]


def normalise_name(name: str) -> str:
    """Lower-case, drop bracketed aliases and punctuation, collapse spaces."""
    text = re.sub(r"[\(\[][^\)\]]*[\)\]]", " ", name.translate(INVISIBLE))
    text = re.sub(r"[^\w\s\uac00-\ud7a3]", " ", text, flags=re.UNICODE)
    return re.sub(r"\s+", " ", text).strip().lower()


def _name_key(name: str) -> tuple[str, ...]:
    return tuple(sorted(normalise_name(name).split()))


def looks_like_name(text: str) -> bool:
    cleaned = text.strip()
    if len(cleaned) < 2:
        return False
    if not re.search(r"[A-Za-z\uac00-\ud7a3]", cleaned):
        return False
    tokens = [t for t in normalise_name(cleaned).split() if t not in _LABEL_FILLER]
    return bool(tokens)


# --------------------------------------------------------------------------
# Month / time parsing
# --------------------------------------------------------------------------


UNNAMED_CLASS = "(unnamed class)"


def infer_month(sheet_name: str) -> int:
    """Return the 1-12 month number encoded in a worksheet name."""
    text = str(sheet_name).strip().lower()
    if not text:
        raise ValueError("The worksheet name is empty.")

    korean = re.search(r"(\d{1,2})\s*월", text)
    if korean:
        month = int(korean.group(1))
        if 1 <= month <= 12:
            return month

    candidates: list[tuple[int, int]] = []
    for index in range(1, 13):
        for label in (calendar.month_name[index], calendar.month_abbr[index]):
            position = text.find(label.lower())
            if position != -1:
                candidates.append((len(label), index))
    if candidates:
        candidates.sort(reverse=True)
        return candidates[0][1]

    numeric = re.findall(r"\d{1,2}", text)
    for value in numeric:
        month = int(value)
        if 1 <= month <= 12:
            return month

    raise ValueError(f"Could not work out a month from the worksheet name {sheet_name!r}.")


def infer_year(sheet_name: str, fallback: int) -> int:
    """The year a worksheet names, or ``fallback`` when it names none.

    Some workbooks keep years apart by sheet -- "2026.Jun", "2025.Dec" -- and
    span two or three of them in one file. Taking the year from the name means
    those sheets land on the right dates instead of every one of them being
    forced into whichever year was picked on the upload form, which silently
    moves a class onto the wrong day of the week.

    Only a full four-digit year counts. A trailing "26" could as easily be a
    day or a class number, and guessing wrong is worse than asking.
    """
    match = re.search(r"(19|20)\d{2}", str(sheet_name))
    return int(match.group(0)) if match else int(fallback)


def _to_time(hour: int, minute: int, meridiem: str | None) -> dt.time:
    if meridiem:
        flag = meridiem.replace(".", "").replace(" ", "").lower()
        if flag.startswith("p") and hour != 12:
            hour += 12
        elif flag.startswith("a") and hour == 12:
            hour = 0
    hour = hour % 24
    return dt.time(hour, minute)


def parse_time_range(text: str) -> tuple[dt.time, dt.time] | None:
    """Parse ``9am-11am`` / ``11.15am-1.15pm`` / ``5pm-6.30pm`` style ranges."""
    match = TIME_RANGE_RE.search(text)
    if not match:
        return None

    start_hour = int(match.group("sh"))
    end_hour = int(match.group("eh"))
    if not (0 < start_hour <= 24 and 0 < end_hour <= 24):
        return None

    start_minute = int(match.group("sm") or 0)
    end_minute = int(match.group("em") or 0)
    start_flag = match.group("sap")
    end_flag = match.group("eap")

    # Academy hours run 08:00-22:00, so infer the missing am/pm marker.
    if start_flag is None and end_flag is None:
        start_flag = "am" if start_hour >= 8 and start_hour != 12 else "pm"
        end_flag = "am" if end_hour >= 8 and end_hour != 12 and end_hour > start_hour else "pm"
    elif start_flag is None:
        start_flag = end_flag
        if _to_time(start_hour, start_minute, start_flag) >= _to_time(
            end_hour, end_minute, end_flag
        ):
            start_flag = "am" if end_flag.lower().startswith("p") else "pm"
    elif end_flag is None:
        end_flag = start_flag
        if _to_time(end_hour, end_minute, end_flag) <= _to_time(
            start_hour, start_minute, start_flag
        ):
            end_flag = "pm" if start_flag.lower().startswith("a") else "am"

    start = _to_time(start_hour, start_minute, start_flag)
    end = _to_time(end_hour, end_minute, end_flag)
    return start, end


def _is_time_line(line: str) -> bool:
    """True when the line is only a time range (not a class name)."""
    match = TIME_RANGE_RE.search(line)
    if not match:
        return False
    remainder = (line[: match.start()] + line[match.end():]).strip()
    remainder = re.sub(r"[\s\-~()\[\].,:]", "", remainder)
    return remainder == ""


def status_from_line(line: str) -> tuple[str | None, str]:
    """Return ``(status, leftover_text)`` for a possible status label line."""
    for status, pattern in STATUS_PATTERNS:
        match = pattern.search(line)
        if match:
            leftover = (line[: match.start()] + " " + line[match.end():]).strip()
            leftover = leftover.strip(" -:/,.()")
            tokens = [
                token
                for token in normalise_name(leftover).split()
                if token not in _LABEL_FILLER
            ]
            return status, leftover if tokens else ""

    # Catch single-word typos such as "Reording" or "Onilne".
    token = re.sub(r"[^a-z]", "", line.lower())
    if len(token) >= 5:
        for status, keyword in ((CANCELLED, "cancelled"), (ONLINE, "online"), (RECORDING, "recording")):
            if difflib.SequenceMatcher(None, token, keyword).ratio() >= 0.82:
                return status, ""

    return None, line


# --------------------------------------------------------------------------
# Cell parsing
# --------------------------------------------------------------------------


def parse_cell(text: str) -> dict[str, Any]:
    """Split a calendar cell into class name, time range and student entries.

    Returns ``{"class_lines", "time_range", "entries", "notes"}`` where each
    entry is ``{"name", "status"}`` and ``status`` is ``None`` for a plain
    roster line.
    """
    lines = _cell_lines(text)
    time_range = None
    time_index = None
    for index, line in enumerate(lines):
        if _is_time_line(line):
            parsed = parse_time_range(line)
            if parsed:
                time_range = parsed
                time_index = index
                break

    class_lines = lines[:time_index] if time_index is not None else []
    body = lines[time_index + 1:] if time_index is not None else lines

    entries: list[dict[str, Any]] = []
    notes: list[str] = []
    current_status: str | None = None

    for line in body:
        status, leftover = status_from_line(line)
        if status is not None:
            current_status = status
            if leftover and looks_like_name(leftover):
                entries.append({"name": leftover, "status": status})
            continue

        if line.startswith("(") and entries:
            entries[-1]["note"] = line.strip("()")
            continue

        if looks_like_name(line):
            entries.append({"name": line, "status": current_status})
        else:
            notes.append(line)

    return {
        "class_lines": class_lines,
        "time_range": time_range,
        "entries": entries,
        "notes": notes,
    }


# --------------------------------------------------------------------------
# Roster matching
# --------------------------------------------------------------------------


class _Roster:
    """Matches names found in status cells back onto the class roster."""

    def __init__(self) -> None:
        self.names: list[str] = []
        self._by_exact: dict[str, int] = {}
        self._by_tokens: dict[tuple[str, ...], int] = {}

    def add(self, name: str) -> int:
        key = normalise_name(name)
        if key in self._by_exact:
            return self._by_exact[key]
        index = len(self.names)
        self.names.append(name)
        self._by_exact[key] = index
        self._by_tokens.setdefault(_name_key(name), index)
        return index

    def match(self, name: str) -> tuple[int | None, str | None]:
        """Return ``(index, warning_kind)``; index is ``None`` when unmatched."""
        key = normalise_name(name)
        if key in self._by_exact:
            return self._by_exact[key], None

        token_key = _name_key(name)
        if token_key in self._by_tokens:
            return self._by_tokens[token_key], "reordered"

        scored = sorted(
            (
                (difflib.SequenceMatcher(None, key, normalise_name(candidate)).ratio(), index)
                for index, candidate in enumerate(self.names)
            ),
            reverse=True,
        )
        if scored:
            best_score, best_index = scored[0]
            runner_up = scored[1][0] if len(scored) > 1 else 0.0
            if best_score >= 0.90 and best_score - runner_up >= 0.05:
                return best_index, "fuzzy"
        return None, None


# --------------------------------------------------------------------------
# Sheet geometry
# --------------------------------------------------------------------------


def _find_header_row(worksheet, max_scan: int = 8) -> tuple[int, dict[int, tuple[int, str | None]]]:
    """Find the row holding ``15(SAT)`` style day labels."""
    best_row = 0
    best_days: dict[int, tuple[int, str | None]] = {}
    limit = min(max_scan, worksheet.max_row or max_scan)

    for row in range(1, limit + 1):
        days: dict[int, tuple[int, str | None]] = {}
        for column in range(1, (worksheet.max_column or 1) + 1):
            value = worksheet.cell(row=row, column=column).value
            if not isinstance(value, str):
                continue
            match = DAY_HEADER_RE.match(value)
            if not match:
                continue
            day = int(match.group("day"))
            dow = (match.group("dow") or "").strip().lower() or None
            if 1 <= day <= 31 and (dow is None or dow in WEEKDAYS):
                days[column] = (day, dow)
        if len(days) > len(best_days):
            best_days = days
            best_row = row
    return best_row, best_days


def _time_axis(worksheet, header_row: int) -> tuple[dict[int, dt.time], set[int]]:
    """Return ``(row -> time)`` and the set of columns used as a time axis."""
    counts: dict[int, int] = {}
    row_times: dict[int, dt.time] = {}

    for column in range(1, (worksheet.max_column or 1) + 1):
        hits = 0
        for row in range(header_row + 1, (worksheet.max_row or 1) + 1):
            value = worksheet.cell(row=row, column=column).value
            if isinstance(value, dt.datetime):
                value = value.time()
            if isinstance(value, dt.time):
                hits += 1
        if hits >= 3:
            counts[column] = hits

    for column in counts:
        for row in range(header_row + 1, (worksheet.max_row or 1) + 1):
            value = worksheet.cell(row=row, column=column).value
            if isinstance(value, dt.datetime):
                value = value.time()
            if isinstance(value, dt.time):
                row_times.setdefault(row, value)

    return row_times, set(counts)


def _fill_key(cell) -> str | None:
    """A comparable identity for a cell's background fill.

    Excel stores a fill three different ways -- a literal RGB value, an index
    into the theme palette, or an indexed legacy colour -- and two cells that
    look identical on screen can be stored differently.  Returns ``None`` for
    an unfilled cell, and a string that only compares equal to a fill of the
    same kind.
    """
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    colour = fill.fgColor
    if colour is None:
        return None
    if colour.type == "rgb" and isinstance(colour.rgb, str):
        return f"rgb:{colour.rgb}:{colour.tint or 0:.3f}"
    if colour.type == "theme" and isinstance(colour.theme, int):
        return f"theme:{colour.theme}:{colour.tint or 0:.3f}"
    if colour.type == "indexed" and isinstance(colour.indexed, int):
        return f"indexed:{colour.indexed}"
    return None


def _merged_lookup(worksheet) -> dict[tuple[int, int], tuple[int, int, int, int]]:
    lookup: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for merged in worksheet.merged_cells.ranges:
        bounds = (merged.min_row, merged.min_col, merged.max_row, merged.max_col)
        lookup[(merged.min_row, merged.min_col)] = bounds
    return lookup


def _row_step_minutes(row_times: dict[int, dt.time]) -> int:
    rows = sorted(row_times)
    deltas = []
    for first, second in zip(rows, rows[1:]):
        if second - first != 1:
            continue
        start = dt.datetime.combine(dt.date.today(), row_times[first])
        end = dt.datetime.combine(dt.date.today(), row_times[second])
        minutes = int((end - start).total_seconds() // 60)
        if 0 < minutes <= 120:
            deltas.append(minutes)
    if not deltas:
        return 30
    return max(set(deltas), key=deltas.count)


def _warning(
    message: str,
    *,
    sheet: str,
    coordinate: str | None = None,
    date: dt.date | None = None,
    cell_text: Any = None,
) -> dict[str, Any]:
    """One warning, kept structured so the screen can point at the cell.

    ``text`` repeats the single line the parser has always produced, so
    anything that only prints warnings keeps working unchanged; the separate
    fields let the import screen show the worksheet, the cell reference and
    what is actually typed in that cell.
    """
    where = f"{sheet}!{coordinate}" if coordinate else sheet
    stamp = f" ({date.isoformat()})" if date else ""
    return {
        "sheet": sheet,
        "coordinate": coordinate,
        "cell": where if coordinate else None,
        "date": date,
        "message": message,
        "cell_text": cell_text if isinstance(cell_text, str) else None,
        "text": f"{where}{stamp}: {message}",
    }


def _resolve_dates(
    days: dict[int, tuple[int, str | None]],
    month: int,
    year: int,
) -> tuple[dict[int, dt.date], list[tuple[int, str]]]:
    """Map header columns to real dates, rolling over into the next month.

    Warnings come back as ``(column, message)`` so the caller can point at
    the day header cell they came from.
    """
    dates: dict[int, dt.date] = {}
    warnings: list[tuple[int, str]] = []
    current_month, current_year = month, year
    previous_day = 0

    for column in sorted(days):
        day, dow = days[column]
        if day < previous_day:  # the sheet spilled into the following month
            current_month += 1
            if current_month > 12:
                current_month, current_year = 1, current_year + 1
        previous_day = day

        last_day = calendar.monthrange(current_year, current_month)[1]
        if day > last_day:
            warnings.append(
                (
                    column,
                    f"day {day} does not exist in "
                    f"{calendar.month_name[current_month]} {current_year}; "
                    "this column was skipped.",
                )
            )
            continue

        value = dt.date(current_year, current_month, day)
        if dow and WEEKDAYS[dow] != value.weekday():
            warnings.append(
                (
                    column,
                    f"{value.isoformat()} is a "
                    f"{calendar.day_abbr[value.weekday()]} but the sheet says "
                    f"{dow.upper()} — check the import year.",
                )
            )
        dates[column] = value

    return dates, warnings


# --------------------------------------------------------------------------
# Sheet parsing
# --------------------------------------------------------------------------


def _parse_sheet(worksheet, month: int, year: int) -> tuple[list[dict[str, Any]], list[str]]:
    sheet_name = worksheet.title
    warnings: list[dict[str, Any]] = []

    header_row, days = _find_header_row(worksheet)
    if not days:
        return [], [
            _warning("no day headers such as '15(SAT)' were found.", sheet=sheet_name)
        ]

    row_times, time_columns = _time_axis(worksheet, header_row)
    if not row_times:
        warnings.append(
            _warning(
                "no time axis column was found; grid times unavailable.",
                sheet=sheet_name,
            )
        )
    step_minutes = _row_step_minutes(row_times)
    merged = _merged_lookup(worksheet)
    dates, date_warnings = _resolve_dates(days, month, year)
    for column, message in date_warnings:
        header_cell = worksheet.cell(row=header_row, column=column)
        warnings.append(
            _warning(
                message,
                sheet=sheet_name,
                coordinate=header_cell.coordinate,
                cell_text=header_cell.value,
            )
        )

    max_row = max(row_times) if row_times else (worksheet.max_row or header_row)
    max_column = worksheet.max_column or 1
    header_columns = sorted(days)

    sessions: list[dict[str, Any]] = []

    for position, anchor in enumerate(header_columns):
        if anchor not in dates:
            continue
        session_date = dates[anchor]

        # The day block runs to the next day header or the next time column.
        block_end = max_column
        if position + 1 < len(header_columns):
            block_end = header_columns[position + 1] - 1
        for column in range(anchor + 1, block_end + 1):
            if column in time_columns:
                block_end = column - 1
                break
        block_columns = [
            column
            for column in range(anchor, block_end + 1)
            if column not in time_columns
        ]

        classes: list[dict[str, Any]] = []
        status_cells: list[tuple[int, int, dict[str, Any], Any, Any]] = []

        for column in block_columns:
            for row in range(header_row + 1, max_row + 1):
                cell = worksheet.cell(row=row, column=column)
                if not isinstance(cell.value, str) or not cell.value.strip():
                    continue
                # Non-anchor cells of a merged range read back as None, so
                # anything with text here is the anchor of its block.
                bounds = merged.get((row, column))
                parsed = parse_cell(cell.value)
                span_end = bounds[2] if bounds else row

                if parsed["time_range"] is not None:
                    start, end = parsed["time_range"]
                    duration = (
                        dt.datetime.combine(dt.date.today(), end)
                        - dt.datetime.combine(dt.date.today(), start)
                    ).total_seconds() / 60
                    rows_needed = max(1, int(round(duration / step_minutes)))
                    classes.append(
                        {
                            "row": row,
                            "column": column,
                            "end_row": max(span_end, row + rows_needed - 1),
                            "parsed": parsed,
                            "coordinate": cell.coordinate,
                            "text": cell.value,
                            "fill": _fill_key(cell),
                        }
                    )
                else:
                    status_cells.append(
                        (row, column, parsed, _fill_key(cell), cell.value)
                    )
                    if not parsed["entries"] and parsed["notes"]:
                        warnings.append(
                            _warning(
                                f"note ignored — {' / '.join(parsed['notes'])}",
                                sheet=sheet_name,
                                coordinate=cell.coordinate,
                                date=session_date,
                                cell_text=cell.value,
                            )
                        )

        # Trim a class span so it cannot swallow the next class in its column.
        for item in classes:
            following = [
                other["row"]
                for other in classes
                if other["column"] == item["column"] and other["row"] > item["row"]
            ]
            if following:
                item["end_row"] = min(item["end_row"], min(following) - 1)

        for item in classes:
            parsed = item["parsed"]
            start, end = parsed["time_range"]
            class_name = " ".join(
                re.sub(r"\s+", " ", line).strip() for line in parsed["class_lines"]
            ).strip()
            session_warnings: list[dict[str, Any]] = []

            if not class_name:
                class_name = UNNAMED_CLASS
                session_warnings.append(
                    _warning(
                        "the class name is missing above the time range — this "
                        "cell starts straight at the time, so the class imports "
                        f"as \u201c{UNNAMED_CLASS}\u201d.",
                        sheet=sheet_name,
                        coordinate=item["coordinate"],
                        date=session_date,
                        cell_text=item["text"],
                    )
                )

            grid_time = row_times.get(item["row"])
            if grid_time and abs(
                (
                    dt.datetime.combine(dt.date.today(), grid_time)
                    - dt.datetime.combine(dt.date.today(), start)
                ).total_seconds()
            ) > 45 * 60:
                session_warnings.append(
                    _warning(
                        f"'{class_name}' starts at {start.strftime('%H:%M')} but "
                        f"sits on the {grid_time.strftime('%H:%M')} grid row.",
                        sheet=sheet_name,
                        coordinate=item["coordinate"],
                        date=session_date,
                        cell_text=item["text"],
                    )
                )

            roster = _Roster()
            attendance: list[dict[str, Any]] = []

            def add_entry(name: str, status: str | None, source: str, note: str | None = None):
                clean = re.sub(r"\s+", " ", name).strip()
                index = roster.add(clean)
                if index == len(attendance):
                    attendance.append(
                        {
                            "student_name": clean,
                            "status": status or ATTENDING,
                            "source": source,
                            "note": note,
                        }
                    )
                    return index, True
                return index, False

            for entry in parsed["entries"]:
                index, created = add_entry(
                    entry["name"], entry["status"], item["coordinate"], entry.get("note")
                )
                if not created and entry["status"]:
                    attendance[index]["status"] = entry["status"]

            for note in parsed["notes"]:
                session_warnings.append(
                    _warning(
                        f"unrecognised line in '{class_name}' — {note}",
                        sheet=sheet_name,
                        coordinate=item["coordinate"],
                        date=session_date,
                        cell_text=item["text"],
                    )
                )

            # Attach the status cells that sit inside this class's rows.
            for row, column, status_parsed, status_fill, status_text in status_cells:
                if not (item["row"] <= row <= item["end_row"]):
                    continue
                if column <= item["column"]:
                    continue
                reachable = [
                    other
                    for other in classes
                    if other["column"] < column and other["row"] <= row <= other["end_row"]
                ]
                # The sheet colours each class block and its status cells the
                # same, so colour decides the owner when it is available and
                # position decides it otherwise.
                coloured = [
                    other
                    for other in reachable
                    if status_fill is not None and other["fill"] == status_fill
                ]
                owner = max(
                    coloured or reachable,
                    key=lambda other: other["column"],
                    default=None,
                )
                if owner is not item:
                    continue
                colour_disagrees = bool(
                    status_fill
                    and item["fill"]
                    and status_fill != item["fill"]
                    and not coloured
                )

                coordinate = f"{get_column_letter(column)}{row}"
                for entry in status_parsed["entries"]:
                    status = entry["status"]
                    if status is None:
                        session_warnings.append(
                            _warning(
                                f"'{entry['name']}' has no Online/Recording/"
                                "Cancelled label; treated as attending.",
                                sheet=sheet_name,
                                coordinate=coordinate,
                                date=session_date,
                                cell_text=status_text,
                            )
                        )
                        status = ATTENDING

                    index, warning_kind = roster.match(entry["name"])
                    if index is None:
                        add_entry(entry["name"], status, coordinate, entry.get("note"))
                        if colour_disagrees:
                            session_warnings.append(
                                _warning(
                                    f"'{entry['name']}' is marked {status} but is "
                                    f"not on the '{class_name}' roster, and this "
                                    "cell is a different colour from that class; "
                                    "it may belong to another class.",
                                    sheet=sheet_name,
                                    coordinate=coordinate,
                                    date=session_date,
                                    cell_text=status_text,
                                )
                            )
                        else:
                            session_warnings.append(
                                _warning(
                                    f"'{entry['name']}' is marked {status} but is "
                                    f"not on the '{class_name}' roster; the cell "
                                    "colour matches the class, so it was added "
                                    "to it.",
                                    sheet=sheet_name,
                                    coordinate=coordinate,
                                    date=session_date,
                                    cell_text=status_text,
                                )
                            )
                        continue

                    if warning_kind:
                        session_warnings.append(
                            _warning(
                                f"'{entry['name']}' matched roster name "
                                f"'{attendance[index]['student_name']}' "
                                f"({warning_kind} spelling).",
                                sheet=sheet_name,
                                coordinate=coordinate,
                                date=session_date,
                                cell_text=status_text,
                            )
                        )
                    previous = attendance[index]["status"]
                    if previous not in (ATTENDING, status):
                        session_warnings.append(
                            _warning(
                                f"'{attendance[index]['student_name']}' is marked "
                                f"both {previous} and {status}; kept {status}.",
                                sheet=sheet_name,
                                coordinate=coordinate,
                                date=session_date,
                                cell_text=status_text,
                            )
                        )
                    attendance[index]["status"] = status
                    if entry.get("note"):
                        attendance[index]["note"] = entry["note"]

                for note in status_parsed["notes"]:
                    session_warnings.append(
                        _warning(
                            f"unrecognised line — {note}",
                            sheet=sheet_name,
                            coordinate=coordinate,
                            date=session_date,
                            cell_text=status_text,
                        )
                    )

            if not attendance:
                session_warnings.append(
                    _warning(
                        f"'{class_name}' has no student names.",
                        sheet=sheet_name,
                        coordinate=item["coordinate"],
                        date=session_date,
                        cell_text=item["text"],
                    )
                )

            sessions.append(
                {
                    "worksheet": sheet_name,
                    "date": session_date,
                    "class_name": class_name,
                    "start_time": start,
                    "end_time": end,
                    "attendance": attendance,
                    "warnings": session_warnings,
                    "cell": f"{sheet_name}!{item['coordinate']}",
                    "coordinate": item["coordinate"],
                    "cell_text": item["text"],
                }
            )

    sessions.sort(key=lambda item: (item["date"], item["start_time"], item["class_name"]))
    return sessions, warnings


# --------------------------------------------------------------------------
# Cross-workbook name canonicalisation
# --------------------------------------------------------------------------


# Punctuation that can only be a leftover from typing several names into one
# cell -- a stray comma, a dangling slash. Never part of a person's name, and
# leaving it attached turns "Woojin" and "Woojin," into two different students.
_EDGE_PUNCTUATION = " -,;/·•"


def _display_clean(name: str) -> str:
    """Tidy a name for display without changing who it refers to."""
    return re.sub(r"\s+", " ", name.translate(INVISIBLE)).strip(_EDGE_PUNCTUATION)


def _bare(name: str) -> str:
    """The name with any bracketed suffix removed, case-folded."""
    text = re.sub(r"\s*[\(\[][^\)\]]*[\)\]]", "", _display_clean(name))
    return text.strip(" -").casefold()


def _suffix(name: str) -> str:
    found = re.findall(r"[\(\[]([^\)\]]*)[\)\]]", _display_clean(name))
    return found[0].strip() if found else ""


_REASON_TEXT = {
    "capitalisation": "differ only by capitalisation",
    "tag": "same name, tagged differently",
    "spelling": "similar spelling",
}

# Shown worst-first: a tag difference is the one most likely to be two real
# people sharing a name (confirmed by the academy -- a bracketed tag is not
# reliably an alias), so it's surfaced ahead of a probably-safe typo.
_REASON_ORDER = {"tag": 0, "spelling": 1, "capitalisation": 2}


def canonicalise_names(sessions: list[dict[str, Any]]) -> dict[str, Any]:
    """Find spellings of the same student across a workbook -- and flag every
    one for a human decision. Nothing is merged automatically here.

    A workbook mixes three kinds of variation, and none of them is safe to
    assume is the same person on its own:

    * Pure case/spacing differences (``Bae Sujin`` vs ``han seoyoung``)
      are usually a typo, but "usually" isn't "always".
    * A bracketed tag (``Seo Yerin`` vs ``Seo Yerin(UWC D)``) is *not*
      assumed to be an alias on one person -- two students can share a name
      and be told apart only by a tag, so a tag difference is flagged, not
      silently stripped and merged.
    * A similar-but-not-identical spelling (``Bae Sujin`` / ``Han
      Seyoung``) is flagged the same way.

    Attendance entries get only cosmetic clean-up here (invisible
    characters, extra whitespace) -- nothing is renamed until a human
    approves a merge via ``apply_review_decisions`` in schedule_backfill.py.
    """
    roster_mates: dict[str, set[str]] = defaultdict(set)
    counts: Counter = Counter()

    for session in sessions:
        anchor = session["coordinate"]
        enrolled = {
            _display_clean(item["student_name"])
            for item in session["attendance"]
            if item["source"] == anchor
        }
        for key in enrolled:
            roster_mates[key] |= enrolled - {key}

    for session in sessions:
        for item in session["attendance"]:
            cleaned = _display_clean(item["student_name"])
            item["raw_name"] = item["student_name"]
            item["tag"] = _suffix(cleaned) or None
            item["student_name"] = cleaned
            counts[cleaned] += 1

    names = sorted(counts)
    reviews: list[dict[str, Any]] = []
    for index, left in enumerate(names):
        for right in names[index + 1:]:
            if left.casefold() == right.casefold():
                reason, ratio = "capitalisation", 1.0
            else:
                bare_left, bare_right = _bare(left), _bare(right)
                if bare_left == bare_right:
                    reason, ratio = "tag", 1.0
                else:
                    ratio = difflib.SequenceMatcher(None, bare_left, bare_right).ratio()
                    reason = "spelling" if ratio >= MERGE_THRESHOLD else None
            if reason is None:
                continue

            detail = _REASON_TEXT[reason]
            if reason == "tag":
                detail += f" ('{_suffix(left) or 'no tag'}' vs '{_suffix(right) or 'no tag'}')"
            detail += (
                " -- they appear on the same roster together"
                if right in roster_mates[left]
                else " -- they never share a class"
            )

            reviews.append(
                {
                    "names": [left, right],
                    "similarity": round(ratio, 3),
                    "reason": reason,
                    "reason_text": detail,
                    "counts": [counts[left], counts[right]],
                }
            )

    reviews.sort(key=lambda item: (_REASON_ORDER[item["reason"]], -item["similarity"]))
    return {"reviews": reviews}


# --------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------


def get_sheet_names(source: Any) -> list[str]:
    workbook = load_workbook(_as_stream(source), data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _summarise(
    sessions: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    sheet_name: str,
    month: int | None,
    year: int,
) -> dict[str, Any]:
    # Counted as the names actually appear, not folded down by
    # normalise_name: that strips bracketed tags, which would report
    # "Seo Yerin" and "Seo Yerin(UWC D)" as one student when the rule is that
    # they are two people. Anything that really is one person typed twice is
    # raised as a review, and only merges once somebody says so.
    unique_students = {
        attendance["student_name"]
        for session in sessions
        for attendance in session["attendance"]
    }
    total_warnings = len(warnings) + sum(len(session["warnings"]) for session in sessions)
    return {
        "sheet_name": sheet_name,
        "month": month,
        "year": year,
        "sessions": sessions,
        "session_count": len(sessions),
        "unique_student_count": len(unique_students),
        "warnings": warnings,
        "warning_count": total_warnings,
        "sheet_summaries": [],
    }


def parse_schedule(source: Any, sheet_name: str, year: int) -> dict[str, Any]:
    """Parse a single worksheet into a preview dictionary."""
    month = infer_month(sheet_name)
    workbook = load_workbook(_as_stream(source), data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"The workbook has no worksheet named {sheet_name!r}.")
        sessions, warnings = _parse_sheet(
            workbook[sheet_name], month, infer_year(sheet_name, int(year))
        )
    finally:
        workbook.close()
    report = canonicalise_names(sessions)
    preview = _summarise(sessions, warnings, sheet_name, month, int(year))
    preview["name_reviews"] = report["reviews"]
    return preview


def parse_workbook(source: Any, sheet_names: Iterable[str], year: int) -> dict[str, Any]:
    """Parse several worksheets and merge them into one preview dictionary."""
    year = int(year)
    workbook = load_workbook(_as_stream(source), data_only=True)
    all_sessions: list[dict[str, Any]] = []
    all_warnings: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []

    try:
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                all_warnings.append(
                    _warning("worksheet not found; skipped.", sheet=sheet_name)
                )
                continue
            try:
                month = infer_month(sheet_name)
            except ValueError as error:
                all_warnings.append(_warning(str(error), sheet=sheet_name))
                continue

            sheet_year = infer_year(sheet_name, year)
            sessions, warnings = _parse_sheet(workbook[sheet_name], month, sheet_year)
            all_sessions.extend(sessions)
            all_warnings.extend(warnings)

            students = {
                attendance["student_name"]
                for session in sessions
                for attendance in session["attendance"]
            }
            summaries.append(
                {
                    "Worksheet": sheet_name,
                    "Month": calendar.month_name[month],
                    "Sessions": len(sessions),
                    "Students": len(students),
                    "Warnings": len(warnings)
                    + sum(len(session["warnings"]) for session in sessions),
                }
            )
    finally:
        workbook.close()

    all_sessions.sort(key=lambda item: (item["date"], item["start_time"], item["class_name"]))

    # Month sheets overlap at the edges (a July sheet often carries 1 August).
    seen: dict[tuple[dt.date, dt.time, str], str] = {}
    for session in all_sessions:
        key = (session["date"], session["start_time"], normalise_name(session["class_name"]))
        first_seen = seen.get(key)
        if first_seen is None:
            seen[key] = session["cell"]
            continue
        all_warnings.append(
            _warning(
                f"'{session['class_name']}' also appears at {first_seen}; the two "
                "worksheets overlap, so import only one of them.",
                sheet=session["worksheet"],
                coordinate=session["coordinate"],
                date=session["date"],
                cell_text=session["cell_text"],
            )
        )

    report = canonicalise_names(all_sessions)
    preview = _summarise(all_sessions, all_warnings, "All worksheets", None, year)
    preview["sheet_summaries"] = summaries
    preview["name_reviews"] = report["reviews"]
    return preview


# --------------------------------------------------------------------------
# Command line helper (handy for checking a workbook outside Streamlit)
# --------------------------------------------------------------------------

if __name__ == "__main__":  # pragma: no cover
    import argparse

    argument_parser = argparse.ArgumentParser(description=__doc__)
    argument_parser.add_argument("workbook")
    argument_parser.add_argument("--year", type=int, default=dt.date.today().year)
    argument_parser.add_argument("--sheet", default=None)
    argument_parser.add_argument("--warnings", action="store_true")
    options = argument_parser.parse_args()

    sheets = get_sheet_names(options.workbook)
    if options.sheet:
        result = parse_schedule(options.workbook, options.sheet, options.year)
    else:
        result = parse_workbook(options.workbook, sheets, options.year)

    print(f"sessions={result['session_count']} students={result['unique_student_count']} "
          f"warnings={result['warning_count']}")
    for session in result["sessions"]:
        counts: dict[str, int] = {}
        for attendance in session["attendance"]:
            counts[attendance["status"]] = counts.get(attendance["status"], 0) + 1
        print(
            f"{session['date']} {session['start_time']:%H:%M}-{session['end_time']:%H:%M} "
            f"| {session['class_name']} | {len(session['attendance'])} students "
            f"| {counts}"
        )
    if options.warnings:
        for warning in result["warnings"]:
            print("!", warning["text"])
        for session in result["sessions"]:
            for warning in session["warnings"]:
                print("!", warning["text"])
