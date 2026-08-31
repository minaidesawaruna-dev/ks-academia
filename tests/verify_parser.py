"""Verify the schedule parser: pure functions, then a real workbook."""
from __future__ import annotations

import datetime as dt
import io
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
PROJECT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT))

from openpyxl import Workbook  # noqa: E402

import schedule_parser as sp  # noqa: E402

results: list[tuple[str, bool, str]] = []


def check(name, fn):
    try:
        results.append((name, True, fn() or ""))
    except Exception as exc:  # noqa: BLE001
        results.append((name, False, f"{type(exc).__name__}: {exc}"))


# ----------------------------------------------------------- pure functions
def t_time_ranges():
    cases = {
        "9am-11am": (dt.time(9, 0), dt.time(11, 0)),
        "9:30am-11:00am": (dt.time(9, 30), dt.time(11, 0)),
        "2pm-3:30pm": (dt.time(14, 0), dt.time(15, 30)),
        "10-11:30am": (dt.time(10, 0), dt.time(11, 30)),
    }
    for text, want in cases.items():
        got = sp.parse_time_range(text)
        assert got == want, f"{text!r} -> {got}, wanted {want}"
    assert sp.parse_time_range("not a time") is None, "accepted nonsense"
    return f"{len(cases)} formats parsed, nonsense rejected"


def t_status_vocabulary():
    cases = {
        "Cancelled": sp.CANCELLED,
        "cancelled online class": sp.CANCELLED,   # cancel wins over online
        "Online": sp.ONLINE,
        "zoom": sp.ONLINE,
        "Recording": sp.RECORDING,
        "취소": sp.CANCELLED,                      # Korean: cancelled
        "온라인": sp.ONLINE,                        # Korean: online
        "녹화": sp.RECORDING,                       # Korean: recording
    }
    for text, want in cases.items():
        got, _ = sp.status_from_line(text)
        assert got == want, f"{text!r} -> {got}, wanted {want}"
    return f"{len(cases)} labels incl. Korean, precedence correct"


def t_month_year_inference():
    for sheet, month in {"Aug": 8, "August": 8, "Sep": 9, "July": 7,
                         "Jan": 1, "Dec": 12}.items():
        got = sp.infer_month(sheet)
        assert got == month, f"{sheet!r} -> {got}, wanted {month}"
    assert sp.infer_year("Aug", 2026) == 2026, "fallback year ignored"
    return "6 month names + year fallback"


def t_name_normalising():
    a = sp.normalise_name("  Kwak   Jun ")
    b = sp.normalise_name("Nam Jihoon")
    assert a == b, f"{a!r} != {b!r}"
    assert sp.looks_like_name("Nam Jihoon"), "rejected a real name"
    # looks_like_name is a low-level predicate: times are removed upstream by
    # _is_time_line and status words by status_from_line, both of which
    # parse_cell applies first. Test those real guards, not a contract this
    # predicate does not claim.
    assert sp._is_time_line("9am-11am"), "time line not recognised"
    assert not sp._is_time_line("Nam Jihoon"), "name mistaken for a time"
    status, leftover = sp.status_from_line("Recording")
    assert status == sp.RECORDING and not leftover.strip(), (status, leftover)
    cell = sp.parse_cell("Recording\nPark Sohee")
    names = [e["name"] for e in cell["entries"]]
    assert names == ["Park Sohee"], f"status leaked into names: {names}"
    assert cell["entries"][0]["status"] == sp.RECORDING, cell["entries"]
    return "times and status words never become student names"


def t_parse_cell():
    cell = sp.parse_cell("G11 Chem HL B\n9am-11am\nNam Jihoon\nOh Minseok")
    assert cell["time_range"] == (dt.time(9, 0), dt.time(11, 0)), cell
    assert cell["class_lines"] == ["G11 Chem HL B"], cell
    names = [e["name"] for e in cell["entries"]]
    assert names == ["Nam Jihoon", "Oh Minseok"], names
    assert all(e["status"] is None for e in cell["entries"]), cell["entries"]
    return f"subject, time and {len(names)} students split out of one cell"


# ------------------------------------------------------------- real workbook
def build_workbook() -> io.BytesIO:
    """A calendar-grid sheet in the shape the teachers actually send."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Aug"

    # Header row: a time-axis header, then day labels merged over two columns
    # (class column + status column), exactly as the docstring describes.
    ws.cell(row=1, column=1, value="Time")
    ws.cell(row=1, column=2, value="3(MON)")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
    ws.cell(row=1, column=4, value="5(WED)")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)

    # 30-minute time axis down column A.
    times = ["9:00", "9:30", "10:00", "10:30", "11:00", "11:30",
             "12:00", "12:30", "13:00", "13:30"]
    for offset, label in enumerate(times):
        ws.cell(row=2 + offset, column=1, value=label)

    # Monday: one class, with a student marked as a recording watcher.
    ws.cell(row=2, column=2,
            value="G11 Chem HL B\n9am-11am\nNam Jihoon\nOh Minseok\nPark Sohee")
    ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=2)
    ws.cell(row=2, column=3, value="Recording\nPark Sohee")
    ws.merge_cells(start_row=2, start_column=3, end_row=5, end_column=3)

    # Wednesday: a different class, one cancellation and one online student.
    ws.cell(row=6, column=4,
            value="G11 Maths SL\n11am-12:30pm\nSeo Yerin\nChoi Doyun\n수민")
    ws.merge_cells(start_row=6, start_column=4, end_row=8, end_column=4)
    ws.cell(row=6, column=5, value="Online\nChoi Doyun\nCancelled\nSeo Yerin")
    ws.merge_cells(start_row=6, start_column=5, end_row=8, end_column=5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


WB = build_workbook().getvalue()


def t_sheet_names():
    names = sp.get_sheet_names(io.BytesIO(WB))
    assert names == ["Aug"], names
    return f"{names}"


def t_parse_schedule():
    out = sp.parse_schedule(io.BytesIO(WB), "Aug", 2026)
    assert out["month"] == 8, out["month"]
    assert out["session_count"] >= 2, f"only {out['session_count']} sessions"
    subjects = {s.get("class_name") or s.get("subject") for s in out["sessions"]}
    assert any("Econs" in str(x) for x in subjects), subjects
    assert any("Maths" in str(x) for x in subjects), subjects
    return (f"{out['session_count']} sessions, "
            f"{out['unique_student_count']} students, "
            f"{out['warning_count']} warnings")


def t_dates_resolved():
    out = sp.parse_schedule(io.BytesIO(WB), "Aug", 2026)
    dates = sorted({s["date"] for s in out["sessions"] if s.get("date")})
    assert dates, "no dates resolved"
    assert all(d.year == 2026 and d.month == 8 for d in dates), dates
    assert dt.date(2026, 8, 3) in dates, f"Monday the 3rd missing: {dates}"
    assert dt.date(2026, 8, 5) in dates, f"Wednesday the 5th missing: {dates}"
    return f"day labels resolved to {[str(d) for d in dates]}"


def t_times_resolved():
    out = sp.parse_schedule(io.BytesIO(WB), "Aug", 2026)
    econs = [s for s in out["sessions"]
             if "Econs" in str(s.get("class_name") or s.get("subject"))][0]
    assert econs.get("start_time") == dt.time(9, 0), econs.get("start_time")
    assert econs.get("end_time") == dt.time(11, 0), econs.get("end_time")
    return "9am-11am read off the class cell"


def t_statuses_applied():
    out = sp.parse_schedule(io.BytesIO(WB), "Aug", 2026)
    found = {}
    for session in out["sessions"]:
        for att in session["attendance"]:
            found[att["student_name"]] = att.get("status")
    assert found.get("Park Sohee") == sp.RECORDING, found
    assert found.get("Choi Doyun") == sp.ONLINE, found
    assert found.get("Seo Yerin") == sp.CANCELLED, found
    assert found.get("Nam Jihoon") == sp.ATTENDING, found
    return ("recording/online/cancelled/attending all assigned to the "
            "right students")


def t_korean_student_parsed():
    out = sp.parse_schedule(io.BytesIO(WB), "Aug", 2026)
    names = {a["student_name"] for s in out["sessions"] for a in s["attendance"]}
    assert "수민" in names, f"Korean student lost: {sorted(names)}"
    return "Korean student name read out of the grid"


def t_parse_workbook_multi():
    out = sp.parse_workbook(io.BytesIO(WB), ["Aug"], 2026)
    assert out["session_count"] >= 2, out["session_count"]
    out2 = sp.parse_workbook(io.BytesIO(WB), ["Aug", "Nope"], 2026)
    assert any("not found" in str(w) for w in out2["warnings"]), out2["warnings"]
    return "multi-sheet merge works; missing sheet warns instead of crashing"


def t_backfill_suggestions():
    import schedule_backfill as sb
    out = sp.parse_schedule(io.BytesIO(WB), "Aug", 2026)
    merges = sb.suggest_subject_merges(out["sessions"])
    matches = sb.suggest_student_matches(out["sessions"])
    assert isinstance(merges, list) and isinstance(matches, list)
    return f"{len(merges)} subject merges, {len(matches)} student matches suggested"


for name, fn in [
    ("time ranges parse", t_time_ranges),
    ("status vocabulary incl. Korean", t_status_vocabulary),
    ("month/year inference", t_month_year_inference),
    ("name normalising", t_name_normalising),
    ("single class cell parses", t_parse_cell),
    ("workbook sheet names", t_sheet_names),
    ("full sheet parses", t_parse_schedule),
    ("day labels -> real dates", t_dates_resolved),
    ("class times resolved", t_times_resolved),
    ("attendance statuses applied", t_statuses_applied),
    ("Korean student parsed", t_korean_student_parsed),
    ("multi-sheet + missing sheet", t_parse_workbook_multi),
    ("backfill suggestions run", t_backfill_suggestions),
]:
    check(name, fn)

width = max(len(n) for n, _, _ in results)
failed = sum(1 for _, ok, _ in results if not ok)
print()
for name, ok, detail in results:
    print(f"  [{'PASS' if ok else 'FAIL'}] {name.ljust(width)}  {detail}")
print(f"\n{len(results) - failed}/{len(results)} passed")
sys.exit(1 if failed else 0)
